# -*- coding: utf-8 -*-
##Pyslvs - Open Source Planar Linkage Mechanism Simulation and Mechanical Synthesis System. 
##Copyright (C) 2016-2018 Yuan Chang
##E-mail: pyslvs@gmail.com
##
##This program is free software; you can redistribute it and/or modify
##it under the terms of the GNU Affero General Public License as published by
##the Free Software Foundation; either version 3 of the License, or
##(at your option) any later version.
##
##This program is distributed in the hope that it will be useful,
##but WITHOUT ANY WARRANTY; without even the implied warranty of
##MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
##GNU Affero General Public License for more details.
##
##You should have received a copy of the GNU Affero General Public License
##along with this program; if not, write to the Free Software
##Foundation, Inc., 59 Temple Place - Suite 330, Boston, MA 02111-1307, USA.

from core.QtModules import *
from core.graphics import PreviewCanvas, replace_by_dict
from core.io import triangle_class
from core.libs import expr_parser
from core.synthesis import CollectionsDialog
import csv
import openpyxl
import pprint
from math import radians
from copy import deepcopy
from re import split as charSplit
'''
'GeneticPrams',
'FireflyPrams',
'defaultSettings',
'DifferentialPrams',
'AlgorithmType',
'Options_show',
'Path_adjust_show',
'Progress_show',
'Series_show',
'PreviewDialog',
'ChartDialog'
'''
from .DimensionalSynthesis_dialog import *
from .Ui_Algorithm import Ui_Form as PathSolving_Form
nan = float('nan')

#Find a name and return the row from the table.
def name_in_table(widget, name: str) -> int:
    for row in range(widget.rowCount()):
        if widget.item(row, 0).text()==name:
            return row

class DimensionalSynthesis(QWidget, PathSolving_Form):
    fixPointRange = pyqtSignal(dict)
    pathChanged = pyqtSignal(dict)
    mergeResult = pyqtSignal(int, tuple)
    
    def __init__(self, parent):
        super(DimensionalSynthesis, self).__init__(parent)
        self.setupUi(self)
        self.mechanismParams = {}
        self.path = {}
        #Just a pointer reference.
        self.collections = parent.CollectionTabPage.CollectionsTriangularIteration.collections
        #Data and functions.
        self.mechanism_data = []
        self.inputFrom = parent.inputFrom
        self.unsaveFunc = parent.workbookNoSave
        self.Settings = deepcopy(defaultSettings)
        self.algorithmParams_default()
        #Canvas
        def get_solutions_func():
            try:
                return replace_by_dict(self.mechanismParams)
            except KeyError:
                return tuple()
        self.PreviewCanvas = PreviewCanvas(get_solutions_func, self)
        self.preview_layout.addWidget(self.PreviewCanvas)
        self.show_solutions.clicked.connect(self.PreviewCanvas.setShowSolutions)
        #Splitter
        self.up_splitter.setSizes([80, 100])
        #Table widget column width.
        self.ground_joints.setColumnWidth(0, 50)
        self.ground_joints.setColumnWidth(1, 80)
        self.ground_joints.setColumnWidth(2, 70)
        self.ground_joints.setColumnWidth(3, 70)
        self.ground_joints.setColumnWidth(4, 80)
        #Default value of algorithm parameters.
        self.type0.clicked.connect(self.algorithmParams_default)
        self.type1.clicked.connect(self.algorithmParams_default)
        self.type2.clicked.connect(self.algorithmParams_default)
        #Signals
        self.Result_list.clicked.connect(self.hasResult)
        self.clear_button.clicked.connect(self.clear_settings)
        self.clear()
    
    def clear(self):
        self.mechanism_data.clear()
        self.Result_list.clear()
        self.clear_settings()
        self.hasResult()
    
    def clear_settings(self):
        self.on_path_clear_clicked()
        self.path.clear()
        self.mechanismParams.clear()
        self.PreviewCanvas.clear()
        self.Settings.clear()
        self.Settings = deepcopy(defaultSettings)
        self.profile_name.setText("No setting")
        self.type2.setChecked(True)
        self.ground_joints.setRowCount(0)
        self.target_points.clear()
        self.Expression.clear()
        self.Link_Expression.clear()
        self.updateRange()
        self.isGenerate()
    
    def loadResults(self, mechanism_data: list):
        for e in mechanism_data:
            self.mechanism_data.append(e)
            self.add_result(e)
    
    def currentPathChanged(self):
        self.pathChanged.emit({
            name: tuple(path)
            for name, path in self.path.items()
        })
        self.isGenerate()
    
    def currentPath(self):
        item = self.target_points.currentItem()
        if item:
            return self.path[item.text()]
        else:
            return []
    
    @pyqtSlot(str)
    def on_target_points_currentTextChanged(self, text=None):
        self.path_list.clear()
        for x, y in self.currentPath():
            self.path_list.addItem("({:.04f}, {:.04f})".format(x, y))
        self.currentPathChanged()
    
    @pyqtSlot()
    def on_path_clear_clicked(self):
        self.currentPath().clear()
        self.path_list.clear()
        self.currentPathChanged()
    
    @pyqtSlot()
    def on_series_clicked(self):
        dlg = Series_show(self)
        dlg.show()
        if dlg.exec_():
            for x, y in dlg.path:
                self.add_point(x, y)
    
    @pyqtSlot()
    def on_path_copy_clicked(self):
        QApplication.clipboard().setText('\n'.join("{},{}".format(x, y) for x, y in self.currentPath()))
    
    @pyqtSlot()
    def on_path_paste_clicked(self):
        self.readPathFromCSV(charSplit(",|\n", QApplication.clipboard().text()))
    
    @pyqtSlot()
    def on_importCSV_clicked(self):
        fileName = self.inputFrom("Path data", ["Text File (*.txt)", "CSV File (*.csv)"])
        if fileName:
            data = []
            with open(fileName, newline=str()) as stream:
                reader = csv.reader(stream, delimiter=' ', quotechar='|')
                for row in reader:
                    data += ' '.join(row).split(',')
            self.readPathFromCSV(data)
    
    def readPathFromCSV(self, data):
        try:
            data = [(round(float(data[i]), 4), round(float(data[i+1]), 4)) for i in range(0, len(data), 2)]
            for e in data:
                self.add_point(e[0], e[1])
        except:
            QMessageBox.warning(self, "File error",
                "Wrong format.\nIt should be look like this:" + "\n0.0,0.0[\\n]"*3
            )
    
    @pyqtSlot()
    def on_importXLSX_clicked(self):
        fileName = self.inputFrom("Excel file", ["Microsoft Office Excel (*.xlsx *.xlsm *.xltx *.xltm)"])
        if fileName:
            wb = openpyxl.load_workbook(fileName)
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            data = []
            #Keep finding until there is no value.
            i = 1
            while True:
                x = ws.cell(row=i, column=1).value
                y = ws.cell(row=i, column=2).value
                if x==None or y==None:
                    break
                try:
                    data.append((round(float(x), 4), round(float(y), 4)))
                except:
                    QMessageBox.warning(self, "File error",
                        "Wrong format.\nThe datasheet seems to including non-digital cell."
                    )
                    break
                i += 1
            for e in data:
                self.add_point(e[0], e[1])
    
    @pyqtSlot()
    def on_pathAdjust_clicked(self):
        dlg = Path_adjust_show(self)
        dlg.show()
        if dlg.exec_():
            self.on_path_clear_clicked()
            for e in dlg.r_path:
                self.add_point(e[0], e[1])
        self.currentPathChanged()
    
    def addPath(self, x, y):
        self.path_list.addItem('({}, {})'.format(x, y))
        self.isGenerate()
    
    def add_point(self, x, y):
        x = round(x, 4)
        y = round(y, 4)
        self.currentPath().append((x, y))
        self.path_list.addItem("({:.04f}, {:.04f})".format(x, y))
        self.currentPathChanged()
    
    @pyqtSlot()
    def on_close_path_clicked(self):
        currentPath = self.currentPath()
        if self.path_list.count() > 1 and currentPath[0]!=currentPath[-1]:
            self.add_point(*currentPath[0])
    
    @pyqtSlot()
    def on_point_up_clicked(self):
        row = self.path_list.currentRow()
        if row > 0 and self.path_list.count() > 1:
            path = self.currentPath()
            path.insert(row-1, (path[row][0], path[row][1]))
            del path[row+1]
            x, y = self.path_list.currentItem().text()[1:-1].split(", ")
            self.path_list.insertItem(row-1, "({}, {})".format(x, y))
            self.path_list.takeItem(row+1)
            self.path_list.setCurrentRow(row-1)
            self.currentPathChanged()
    
    @pyqtSlot()
    def on_point_down_clicked(self):
        row = self.path_list.currentRow()
        if row < self.path_list.count()-1 and self.path_list.count() > 1:
            path = self.currentPath()
            path.insert(row+2, (path[row][0], path[row][1]))
            del path[row]
            x, y = self.path_list.currentItem().text()[1:-1].split(", ")
            self.path_list.insertItem(row+2, "({}, {})".format(x, y))
            self.path_list.takeItem(row)
            self.path_list.setCurrentRow(row+1)
            self.currentPathChanged()
    
    @pyqtSlot()
    def on_point_delete_clicked(self):
        row = self.path_list.currentRow()
        if row>-1:
            del self.currentPath()[row]
            self.path_list.takeItem(row)
            self.currentPathChanged()
    
    def isGenerate(self):
        self.pointNum.setText(
            "<html><head/><body><p><span style=\"font-size:12pt; color:#00aa00;\">" +
            str(self.path_list.count()) +
            "</span></p></body></html>"
        )
        n = bool(self.mechanismParams) and (self.path_list.count() > 1)
        self.pathAdjust.setEnabled(n)
        self.generate_button.setEnabled(n)
    
    @pyqtSlot()
    def on_generate_button_clicked(self):
        #Check if the number of target points are same.
        leng = -1
        for path in self.path.values():
            if leng<0:
                leng = len(path)
            if len(path)!=leng:
                QMessageBox.warning(self, "Target Error",
                    "The length of target paths should be the same."
                )
                return
        #Get the algorithm type.
        if self.type0.isChecked():
            type_num = AlgorithmType.RGA
        elif self.type1.isChecked():
            type_num = AlgorithmType.Firefly
        elif self.type2.isChecked():
            type_num = AlgorithmType.DE
        #Deep copy it so the pointer will not the same.
        mechanismParams = deepcopy(self.mechanismParams)
        mechanismParams['Target'] = deepcopy(self.path)
        for key in ('Driver', 'Follower'):
            for name in mechanismParams[key]:
                row = name_in_table(self.ground_joints, name)
                mechanismParams[key][name] = (
                    self.ground_joints.cellWidget(row, 2).value(),
                    self.ground_joints.cellWidget(row, 3).value(),
                    self.ground_joints.cellWidget(row, 4).value()
                )
        for name in ['IMax', 'IMin', 'LMax', 'LMin', 'FMax', 'FMin', 'AMax', 'AMin']:
            mechanismParams[name] = self.Settings[name]
        setting = {'report': self.Settings['report']}
        if 'maxGen' in self.Settings:
            setting['maxGen'] = self.Settings['maxGen']
        elif 'minFit' in self.Settings:
            setting['minFit'] = self.Settings['minFit']
        elif 'maxTime' in self.Settings:
            setting['maxTime'] = self.Settings['maxTime']
        setting.update(self.Settings['algorithmPrams'])
        #Start progress dialog.
        dlg = Progress_show(
            type_num,
            mechanismParams,
            setting,
            self.portText.text() if self.has_zmq.isChecked() else None,
            self
        )
        dlg.show()
        if dlg.exec_():
            for m in dlg.mechanisms:
                self.mechanism_data.append(m)
                self.add_result(m)
            self.setTime(dlg.time_spand)
            self.unsaveFunc()
            QMessageBox.information(self, "Dimensional Synthesis", "Your tasks is all completed.", QMessageBox.Ok, QMessageBox.Ok)
            print("Finished.")
    
    def setTime(self, time_spand):
        sec = round(time_spand%60, 2)
        mins = int(time_spand/60)
        self.timeShow.setText(
            "<html><head/><body><p><span style=\"font-size:16pt\">"+
            "{}[min] {:.02f}[s]".format(mins, sec)+
            "</span></p></body></html>"
        )
    
    #Add result items, except add to the list.
    def add_result(self, result: dict):
        item = QListWidgetItem(result['Algorithm'])
        interrupt = result['interrupted']
        if interrupt=='False':
            item.setIcon(QIcon(QPixmap(":/icons/task-completed.png")))
        elif interrupt=='N/A':
            item.setIcon(QIcon(QPixmap(":/icons/question-mark.png")))
        else:
            item.setIcon(QIcon(QPixmap(":/icons/interrupted.png")))
        text = "{} ({})".format(
            result['Algorithm'],
            "No interrupt." if interrupt=='False' else "Interrupt at {}".format(interrupt)
        )
        if interrupt=='N/A':
            text += "\n※Completeness is not clear."
        item.setToolTip(text)
        self.Result_list.addItem(item)
    
    @pyqtSlot()
    def on_deleteButton_clicked(self):
        row = self.Result_list.currentRow()
        if row>-1:
            reply = QMessageBox.question(self, "Delete", "Delete this result from list?",
                (QMessageBox.Apply | QMessageBox.Cancel), QMessageBox.Apply)
            if reply==QMessageBox.Apply:
                del self.mechanism_data[row]
                self.Result_list.takeItem(row)
                self.unsaveFunc()
                self.hasResult()
    
    def hasResult(self, p0=None):
        for button in [
            self.mergeButton,
            self.deleteButton,
            self.Result_load_settings,
            self.Result_chart,
            self.Result_clipboard
        ]:
            button.setEnabled(self.Result_list.currentRow()>-1)
    
    @pyqtSlot(QModelIndex)
    def on_Result_list_doubleClicked(self, index):
        row = self.Result_list.currentRow()
        if row>-1:
            dlg = PreviewDialog(self.mechanism_data[row], self.get_path(row), self)
            dlg.show()
            dlg.exec_()
    
    @pyqtSlot()
    def on_mergeButton_clicked(self):
        row = self.Result_list.currentRow()
        if row>-1:
            reply = QMessageBox.question(self, "Merge", "Merge this result to your canvas?",
                (QMessageBox.Apply | QMessageBox.Cancel), QMessageBox.Apply)
            if reply==QMessageBox.Apply:
                self.mergeResult.emit(row, self.get_path(row))
    
    def get_path(self, row):
        Result = self.mechanism_data[row]
        expr_angles, expr_links, expr_points = triangle_class(Result['Expression'])
        if len(expr_angles)>1:
            return tuple()
        '''
        expr_angles: ('a0', ...)
        expr_links: ('L0', 'L1', 'L2', ...)
        expr_points: ('A', 'B', 'C', 'D', 'E', ...)
        '''
        Paths = tuple([] for i in range(len(expr_points)))
        for a in range(360+1):
            data_dict = {e: Result[e] for e in expr_links}
            data_dict.update({e: Result[e] for e in Result['Driver']})
            data_dict.update({e: Result[e] for e in Result['Follower']})
            data_dict.update({expr_angles[0]: radians(a)})
            expr_parser(Result['Expression'], data_dict)
            for i, e in enumerate(expr_points):
                x, y = data_dict[e]
                if x!=nan:
                    Paths[i].append((x, y))
        return tuple(tuple(path) if len(set(path))>1 else tuple() for path in Paths)
    
    @pyqtSlot()
    def on_Result_chart_clicked(self):
        dlg = ChartDialog("Convergence Value", self.mechanism_data, self)
        dlg.show()
    
    @pyqtSlot()
    def on_Result_clipboard_clicked(self):
        QApplication.clipboard().setText(pprint.pformat(self.mechanism_data[self.Result_list.currentRow()]))
    
    @pyqtSlot()
    def on_save_button_clicked(self):
        if self.mechanismParams:
            name, ok = QInputDialog.getText(self, "Profile name", "Please enter the profile name:")
            if ok:
                i = 0
                while (name not in self.collections) and (not name):
                    name = "Structure_{}".format(i)
                mechanismParams = deepcopy(self.mechanismParams)
                for key in [
                    'Driver',
                    'Follower',
                    'Target'
                ]:
                    for name in mechanismParams[key]:
                        mechanismParams[key][name] = None
                self.collections[name] = mechanismParams
                self.unsaveFunc()
    
    @pyqtSlot()
    def on_load_profile_clicked(self):
        dlg = CollectionsDialog(self)
        dlg.show()
        if dlg.exec_():
            self.clear_settings()
            self.mechanismParams = dlg.mechanismParams
            self.profile_name.setText(dlg.name_loaded)
            self.Expression.setText(self.mechanismParams['Expression'])
            self.Link_Expression.setText(self.mechanismParams['Link_Expression'])
            self.set_ground_joints()
            self.isGenerate()
    
    def set_ground_joints(self):
        params = self.mechanismParams
        self.path.clear()
        self.target_points.clear()
        for name in sorted(params['Target']):
            self.target_points.addItem(name)
            path = params['Target'][name]
            if path:
                self.path[name] = path.copy()
            else:
                self.path[name] = []
        if self.target_points.count():
            self.target_points.setCurrentRow(0)
        gj = {}
        for key in ['Driver', 'Follower']:
            gj.update(params[key])
        self.ground_joints.setRowCount(0)
        self.ground_joints.setRowCount(len(gj))
        def spinbox(v, prefix=False):
            s = QDoubleSpinBox(self)
            s.setMinimum(-1000000.0)
            s.setMaximum(1000000.0)
            s.setSingleStep(10.0)
            s.setValue(v)
            if prefix:
                s.setPrefix("±")
            return s
        nd = {k: int(v.replace('P', '')) for k, v in params['name_dict'].items()}
        for row, name in enumerate(sorted(gj)):
            coord = gj[name]
            self.ground_joints.setItem(row, 0, QTableWidgetItem(name))
            self.ground_joints.setItem(row, 1,
                QTableWidgetItem('Driver' if name in params['Driver'] else 'Follower')
            )
            self.ground_joints.setCellWidget(row, 2,
                spinbox(coord[0] if coord else params['pos'][nd[name]][0])
            )
            self.ground_joints.setCellWidget(row, 3,
                spinbox(coord[1] if coord else params['pos'][nd[name]][1])
            )
            self.ground_joints.setCellWidget(row, 4,
                spinbox(coord[2] if coord else 50., True)
            )
        for row in range(self.ground_joints.rowCount()):
            for column in range(2, 5):
                self.ground_joints.cellWidget(row, column).valueChanged.connect(self.updateRange)
        self.updateRange()
        self.PreviewCanvas.from_profile(self.mechanismParams)
    
    @pyqtSlot()
    def on_Result_load_settings_clicked(self):
        self.hasResult()
        row = self.Result_list.currentRow()
        if row>-1:
            self.clear_settings()
            Result = self.mechanism_data[row]
            if Result['Algorithm'] == str(AlgorithmType.RGA):
                self.type0.setChecked(True)
            elif Result['Algorithm'] == str(AlgorithmType.Firefly):
                self.type1.setChecked(True)
            elif Result['Algorithm'] == str(AlgorithmType.DE):
                self.type2.setChecked(True)
            self.profile_name.setText("External setting")
            #External setting.
            self.Expression.setText(Result['Expression'])
            self.Link_Expression.setText(Result['Link_Expression'])
            #Copy to mechanism params.
            self.mechanismParams.clear()
            for key in [
                'Driver',
                'Follower',
                'Target'
            ]:
                self.mechanismParams[key] = Result[key].copy()
            for key in [
                'Link_Expression',
                'Expression',
                'constraint',
                'Graph',
                'name_dict',
                'pos',
                'cus',
                'same'
            ]:
                self.mechanismParams[key] = Result[key]
            self.set_ground_joints()
            self.setTime(Result['time'])
            settings = Result['settings']
            self.Settings = {
                'report': settings['report'],
                'IMax': Result['IMax'], 'IMin': Result['IMin'],
                'LMax': Result['LMax'], 'LMin': Result['LMin'],
                'FMax': Result['FMax'], 'FMin': Result['FMin'],
                'AMax': Result['AMax'], 'AMin': Result['AMin']
            }
            if 'maxGen' in settings:
                self.Settings['maxGen'] = settings['maxGen']
            elif 'minFit' in settings:
                self.Settings['minFit'] = settings['minFit']
            elif 'maxTime' in setting:
                self.Settings['maxTime'] = setting['maxTime']
            algorithmPrams = settings.copy()
            del algorithmPrams['report']
            self.Settings['algorithmPrams'] = algorithmPrams
    
    def algorithmParams_default(self):
        if self.type0.isChecked():
            self.Settings['algorithmPrams'] = GeneticPrams.copy()
        elif self.type1.isChecked():
            self.Settings['algorithmPrams'] = FireflyPrams.copy()
        elif self.type2.isChecked():
            self.Settings['algorithmPrams'] = DifferentialPrams.copy()
    
    #Get the settings from advance dialog.
    @pyqtSlot()
    def on_advanceButton_clicked(self):
        if self.type0.isChecked():
            type_num = AlgorithmType.RGA
        elif self.type1.isChecked():
            type_num = AlgorithmType.Firefly
        elif self.type2.isChecked():
            type_num = AlgorithmType.DE
        dlg = Options_show(type_num, self.Settings)
        dlg.show()
        if dlg.exec_():
            tablePL = lambda row: dlg.PLTable.cellWidget(row, 1).value()
            self.Settings = {
                'report': dlg.report.value(),
                'IMax': tablePL(0), 'IMin': tablePL(1),
                'LMax': tablePL(2), 'LMin': tablePL(3),
                'FMax': tablePL(4), 'FMin': tablePL(5),
                'AMax': tablePL(6), 'AMin': tablePL(7)
            }
            if dlg.maxGen_option.isChecked():
                self.Settings['maxGen'] = dlg.maxGen.value()
            elif dlg.minFit_option.isChecked():
                self.Settings['minFit'] = dlg.minFit.value()
            elif dlg.maxTime_option.isChecked():
                #Three spinbox value translate to second.
                self.Settings['maxTime'] = (
                    dlg.maxTime_h.value() * 3600 +
                    dlg.maxTime_m.value() * 60 +
                    dlg.maxTime_s.value()
                )
            tableAP = lambda row: dlg.APTable.cellWidget(row, 1).value()
            popSize = dlg.popSize.value()
            if type_num == AlgorithmType.RGA:
                self.Settings['algorithmPrams'] = {
                    'nPop': popSize,
                    'pCross': tableAP(0),
                    'pMute': tableAP(1),
                    'pWin': tableAP(2),
                    'bDelta': tableAP(3)
                }
            elif type_num == AlgorithmType.Firefly:
                self.Settings['algorithmPrams'] = {
                    'n': popSize,
                    'alpha': tableAP(0),
                    'betaMin': tableAP(1),
                    'gamma': tableAP(2),
                    'beta0': tableAP(3)
                }
            elif type_num == AlgorithmType.DE:
                self.Settings['algorithmPrams'] = {
                    'NP': popSize,
                    'strategy': tableAP(0),
                    'F': tableAP(1),
                    'CR': tableAP(2)
                }
    
    @pyqtSlot(float)
    def updateRange(self, p0=None):
        def t(x, y):
            item = self.ground_joints.item(x, y)
            if item:
                return item.text()
            else:
                return self.ground_joints.cellWidget(x, y).value()
        self.fixPointRange.emit({
            t(row, 0): (t(row, 2), t(row, 3), t(row, 4))
            for row in range(self.ground_joints.rowCount())
        })
    
    @pyqtSlot()
    def on_Expression_copy_clicked(self):
        text = self.Expression.text()
        if text:
            QApplication.clipboard().setText(text)
    
    @pyqtSlot()
    def on_Link_Expression_copy_clicked(self):
        text = self.Link_Expression.text()
        if text:
            QApplication.clipboard().setText(text)
